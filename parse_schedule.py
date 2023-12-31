# by 11Б
import openpyxl
import re
from openpyxl.utils.cell import coordinate_to_tuple

def getMergedCellVal(sheet, cell):
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
    return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng) != 0 else cell.value


def parse(need_class, file):
    if file=="":
        return None
    need_class=need_class.replace(" ", "")
    current_file = openpyxl.load_workbook(file)
    table = current_file.active
    classes = []
    times = []
    cell = None
    merged = ''
    for i in str(table.merged_cells):
        if i != ':':
            merged += i
            continue
        break
    if merged:
        row, col = coordinate_to_tuple(merged)
        cell = table.cell(row=row, column=col)
        merged_value = getMergedCellVal(table, cell)

    def get_shedule(need_class):
        for row in table.iter_rows():  # type: ignore
            for cell in row:
                if cell.value == need_class:
                    row_num = cell.row
                    col_num = cell.column
                    for i in range(row_num + 1, row_num + 12):
                        if re.fullmatch(r'\d{2}:\d{2} - \d{2}:\d{2}',
                                        str(table.cell(row=i, column=2).value)) is not None:  # type: ignore
                            cell = table.cell(row=i, column=col_num)
                            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                                classes.append(merged_value)
                            else:
                                classes.append(str(cell.value).replace("\n", ""))  # type: ignore
                            if classes[-1]=="None":
                                classes.pop()
                                continue
                            cell = table.cell(row=i, column=2)
                            times.append(cell.value)  # type: ignore
                        else:
                            break
        shedule = list(map(lambda x, y: [x, y], times, classes))

        def shedule_split(shedule):
            chto = ["гебраических зада","ометричиских зада","физичес","Всеобщая история","изическая культу", "ометрическ", "Изобразительное","Родная литература","Физика в задачах","Родной язык","История России","Всеобщая история","Биология в задачах","альный проект","именение математических","нглийский", "ероятность"]
            nachto = ["Математика","Геометрия","Физика","История","Физ-ра", "Геометрия", "ИЗО","Литература","Физика","Русский язык","История","История","Биология с эксп.","Индив. проект","Математика","Англ. яз","Вероятн. и стат."]
            for i in shedule:
                if i[-1] is not None:
                    index=1000
                    indexmx=0
                    cnt=0
                    for j in ["1","2","3","4","5","6","7","8","9","0","/","Геол.музей","Б.спорт.зал","М.спорт.зал","Акт.зал","Ист.музей"]:
                        if j in i[-1]:
                            index=min(index,i[-1].index(j))
                            indexmx=max(indexmx,i[-1].rindex(j))
                            cnt+=i[-1].count(j)
                    if indexmx-index!=cnt-1:
                        num=""
                    else:
                        num=i[-1][index:]
                        i[-1] = i[-1][:index]
                        for j,k in enumerate(chto):
                            if k in i[-1]:
                                i[-1]=nachto[j]
                    i.append(num)
            return shedule

        return shedule_split(shedule)
    
    return get_shedule(need_class)